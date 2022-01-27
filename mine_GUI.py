import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import random
from tkinter import *
from itertools import count
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
import pandas as pd
from functools import partial  
from tkinter import ttk
import openpyxl
import pathlib
# from csv import DictWriter
import os 



# initalise the tkinter GUI
root = tk.Tk()
root.title('') 
# root.configure(background="#232939")
root.geometry("1100x650") # set the root dimensions
root.pack_propagate(False) # tells the root to not let the widgets inside it determine its size.
root.resizable(0, 0) # makes the root window fixed in size.




# Frame for TreeView
frame1 = tk.LabelFrame(root, text="Excel Data")
frame1.place(height=320, width=600, rely=0, relx=0.01)



# Frame for open file dialog    
file_frame = tk.LabelFrame(root, text="Open File(.xlsx )", bg="gray")
file_frame.place(height=70, width=330, rely=0.49, relx=0.625)


# Buttons
button1 = tk.Button(file_frame, text="Browse A File", activebackground = "pink", activeforeground = "blue", command=lambda: File_dialog())
button1.place(rely=0.43, relx=0.30)

button2 = tk.Button(file_frame, text="Load File", command=lambda: Load_excel_data())
button2.place(rely=0.43, relx=0.06)

# The file/file path text
label_file = ttk.Label(file_frame, text="No File Selected")
label_file.place( rely=0, relx=0.01)


store_data = tk.LabelFrame(root, text="Store Data")
store_data.place(height=320, width=460, rely=0, relx=0.57)



northing_var = tk.StringVar()  
easting_var = tk.StringVar()
grade_var = tk.StringVar()
oredepth_var = tk.StringVar()
overburden_var = tk.StringVar()



def Submit():
   file = pathlib.Path("record.xlsx")
   
   
   if file.exists():
        file = openpyxl.load.Workbook("record.xlsx")
        sheet = file["people"]
        sheet.cell(column=1, row=sheet.max_row + 1, value=northing_var.get())
        sheet.cell(column=2, row=sheet.max_row, value=easting_var.get())
        sheet.cell(column=3, row=sheet.max_row, value=grade_var.get())
        file.save("record.xlsx")
   else:
        file = openpyxl.Workbook()
        sheet = file["Sheet"]
        sheet.title = "people"
        sheet["A1"] = "Northing"
        sheet["B1"] = "Easting"
        sheet["C1"] = "Grading"
        sheet.cell(column=1, row=sheet.max_row + 1, value=northing_var.get())
        sheet.cell(column=2, row=sheet.max_row, value=easting_var.get())
        sheet.cell(column=3, row=sheet.max_row, value=grade_var.get())
        file.save("record.xlsx")

def check(*args):
    errors = 0
    if northing_var.get()=="":
        messagebox.showerror("Empty entry!")
        errors += 1
    if easting_var.get()=="" and errors ==0:
        messagebox.showerror("Empty entry!")
        errors += 1
    if errors ==0:
        Submit()


    if grade_var.get()=="" and errors ==0:
        messagebox.showerror("Empty entry!")
        errors += 1





# clear data from entry

def clearEntryInputnorth():
    northing_ety.delete(0,END)
def clearEntryInputeast():
    easting_ety.delete(0,END)
def clearEntryInputgrade():
    grade_ety.delete(0, END)
def clearEntryInputoredepth():
    oredepth_ety.delete(0, END)
def clearEntryInputoverburden():
    overburden_ety.delete(0, END)


# Box for data storing phase
boxing = tk.LabelFrame(store_data, text="Function button")
boxing.place(height=200, width=305, rely=0.6, relx=0.18)


boring = tk.LabelFrame(store_data, text="Ore Blocking Data")
boring.place(height=90, width=442, rely=0.08, relx=0.01)




northing = Label(boring, text="Northing", font="pristina")
northing.place(height=20, width=80, rely=0.1, relx=0.01)

easting = Label(boring, text="Easting", font="pristina")
easting.place(height=20, width=80, rely=0.1, relx=0.21)

grade = Label(boring, text="Grading", font="pristina")
grade.place(height=20, width=80, rely=0.1, relx=0.41)

ore_depth = Label(boring, text="Ore_depth", font="pristina")
ore_depth.place(height=20, width=80, rely=0.1, relx=0.61)

over_burden = Label(boring, text="Over_burden", font="pristina")
over_burden.place(height=20, width=80, rely=0.1, relx=0.81)



cleartools = tk.LabelFrame(store_data, text="Clear button")
cleartools.place(height=80, width=400, rely=0.31, relx=0.05)


northing_ety = ttk.Entry(boring, textvariable=northing_var, width=14)
northing_ety.grid(columnspan=2)
northing_ety.place(height=20, width=80, rely=0.34, relx=0.01)

easting_ety = ttk.Entry(boring, textvariable=easting_var, width=14)
easting_ety.grid(columnspan=2)
easting_ety.place(height=20, width=80, rely=0.34, relx=0.21)

grade_ety = ttk.Entry(boring, textvariable=grade_var, width=14)
grade_ety.grid(columnspan=2)
grade_ety.place(height=20, width=80, rely=0.34, relx=0.41)

oredepth_ety = ttk.Entry(boring, textvariable=oredepth_var, width=14)
oredepth_ety.grid(columnspan=2)
oredepth_ety.place(height=20, width=80, rely=0.34, relx=0.61)

overburden_ety = ttk.Entry(boring, textvariable=overburden_var, width=14)
overburden_ety.grid(columnspan=2)
overburden_ety.place(height=20, width=80, rely=0.34, relx=0.81)


    


 # clear input data 
NL = Label(cleartools, text="NRH", font="pristina")
NL.place(height=20, width=60, rely=0.2, relx=0.02)
Button(cleartools, text="Clear", fg="white", bg="black", font="chiller", command=clearEntryInputnorth).place(height=20, width=60, rely=0.55, relx=0.02)

EL = Label(cleartools, text="EST", font="pristina")
EL.place(height=20, width=60, rely=0.2, relx=0.21)
Button(cleartools, text="Clear", fg="white", bg="black", font="chiller", command=clearEntryInputeast).place(height=20, width=60, rely=0.55, relx=0.21)

GL = Label(cleartools, text="GRD", font="pristina")
GL.place(height=20, width=60, rely=0.2, relx=0.4)
Button(cleartools, text="Clear", fg="white", bg="black", font="chiller", command=clearEntryInputgrade).place(height=20, width=60, rely=0.55, relx=0.4)

ODL = Label(cleartools, text="ODP", font="pristina")
ODL.place(height=20, width=60, rely=0.2, relx=0.61)
Button(cleartools, text="Clear", fg="white", bg="black", font="chiller", command=clearEntryInputoredepth).place(height=20, width=60, rely=0.55, relx=0.61)

OVB = Label(cleartools, text="OVB", font="pristina")
OVB.place(height=20, width=60, rely=0.2, relx=0.8)
Button(cleartools, text="Clear", fg="white", bg="black", font="chiller", command=clearEntryInputoverburden).place(height=20, width=60, rely=0.55, relx=0.8)

    

# Function button
functools = tk.LabelFrame(store_data, text="Function button")
functools.place(height=60, width=305, rely=0.6, relx=0.18)
button=Button(functools, text="Submit", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=20, width=80, rely=0.2, relx=0.05)
# Button for closing
exit_button = tk.Button(store_data, text="Exit", fg="red", bg="white", command=root.destroy ).place(rely=0, relx=0.92) 



cutoffgrade_var = tk.StringVar()  
oredep_var = tk.StringVar()
Overburdendep_var = tk.StringVar()

length_var = tk.StringVar()
breath_var = tk.StringVar()
height_var = tk.StringVar()

# volume computation 
volume_comp = tk.LabelFrame(root, text="Volume Computation")
volume_comp.place(height=320, width=600, rely=0.5, relx=0.01)

cutdata = tk.LabelFrame(volume_comp)
cutdata.place(height=90, width=305, rely=0.01, relx=0.03)

cutoff_grade= Label(cutdata, text="Cut of Grade")
cutoff_grade.place(height=20, width=80, rely=0.02, relx=0.01)

ore_dep= Label(cutdata, text="Ore Depth")
ore_dep.place(height=20, width=80, rely=0.32, relx=0.01)


Overburden_depth= Label(cutdata, text="OverBurden Depth")
Overburden_depth.place(height=20, width=105, rely=0.62, relx=0.01)

length0= Label(cutdata, text="Length")
length0.place(height=20, width=45, rely=0.1, relx=0.65)

breath0= Label(cutdata, text="Breath")
breath0.place(height=20, width=45, rely=0.42, relx=0.65)

height0= Label(cutdata, text="height")
height0.place(height=20, width=45, rely=0.72, relx=0.65)



cutoff_grade_ety = ttk.Entry(cutdata, textvariable=cutoffgrade_var, width=14)
cutoff_grade_ety.grid(columnspan=2)
cutoff_grade_ety.place(height=20, width=70, rely=0.02, relx=0.3)

ore_dep_ety = ttk.Entry(cutdata, textvariable=oredep_var, width=14)
ore_dep_ety.grid(columnspan=2)
ore_dep_ety.place(height=20, width=70, rely=0.32, relx=0.3)

Overburdendep_ety = ttk.Entry(cutdata, textvariable=Overburdendep_var, width=14)
Overburdendep_ety.grid(columnspan=2)
Overburdendep_ety.place(height=20, width=70, rely=0.61, relx=0.4)


length_ety = ttk.Entry(cutdata, textvariable=length_var, width=14)
length_ety.grid(columnspan=2)
length_ety.place(height=20, width=60, rely=0.1, relx=0.79)

breath_ety = ttk.Entry(cutdata, textvariable=breath_var, width=14)
breath_ety.grid(columnspan=2)
breath_ety.place(height=20, width=60, rely=0.42, relx=0.79)

height_ety = ttk.Entry(cutdata, textvariable=height_var, width=14)
height_ety.grid(columnspan=2)
height_ety.place(height=20, width=60, rely=0.72, relx=0.79)


# length, breath and height computation
    
volumeframe = tk.LabelFrame(volume_comp)
volumeframe.place(height=110, width=315, rely=0.32, relx=0.02)


    
loader = tk.LabelFrame(volume_comp)
loader.place(height=40, width=305, rely=0.7, relx=0.02)
cutdata_submit = Button(loader, text="Submit", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=20, width=60, rely=0.24, relx=0.01)
area_result =Button(loader, text="Area", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=20, width=60, rely=0.24, relx=0.34)
volume_result=Button(loader, text="Volume", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=20, width=60, rely=0.24, relx=0.68)

loader = tk.LabelFrame(volume_comp)
loader.place(height=40, width=305, rely=0.85, relx=0.02)
cutdata_submit = Button(loader, text="Ave Thickness", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=21, width=95, rely=0.24, relx=0.01)
area_result =Button(loader, text="Ave Grade", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=21, width=95, rely=0.24, relx=0.34)
volume_result=Button(loader, text="Total Reserve", activebackground='#31ce61', borderwidth='0', cursor='hand2', fg="white", activeforeground='white', bg="black", font="pristina", command=check).place(height=21, width=95, rely=0.24, relx=0.68)





# volume computation 
comp_output = tk.LabelFrame(volume_comp)
comp_output.place(height=300, width=265, rely=0, relx=0.55)


output_title= Label(comp_output, text="Output Result", font="algerian")
output_title.place(height=20, width=160, rely=0.01, relx=0.21)

# volume and Area output
area_lab= Label(comp_output, text="Area")
area_lab.place(height=20, width=35, rely=0.09, relx=0.08)

area_ety = ttk.Entry(comp_output, textvariable=northing_var, width=14)
area_ety.grid(columnspan=2)
area_ety.place(height=20, width=80, rely=0.09, relx=0.37)

volume_lab= Label(comp_output, text="Volume")
volume_lab.place(height=20, width=45, rely=0.21, relx=0.08)

volume_ety = ttk.Entry(comp_output, textvariable=northing_var, width=14)
volume_ety.grid(columnspan=2)
volume_ety.place(height=20, width=80, rely=0.21, relx=0.37)

avethick_lab= Label(comp_output, text="Aveg-Thickness")
avethick_lab.place(height=20, width=95, rely=0.34, relx=0.00)

avethick_ety = ttk.Entry(comp_output, textvariable=northing_var, width=14)
avethick_ety.grid(columnspan=2)
avethick_ety.place(height=20, width=80, rely=0.34, relx=0.37)

avegrade_lab= Label(comp_output, text="Aveg-Grade")
avegrade_lab.place(height=20, width=80, rely=0.47, relx=0.00)

avegrade_ety = ttk.Entry(comp_output, textvariable=northing_var, width=14)
avegrade_ety.grid(columnspan=2)
avegrade_ety.place(height=20, width=80, rely=0.47, relx=0.37)

totalreserve_lab= Label(comp_output, text="Total-Reserve")
totalreserve_lab.place(height=20, width=80, rely=0.58, relx=0.00)

totalreserve_ety = ttk.Entry(comp_output, textvariable=northing_var, width=14)
totalreserve_ety.grid(columnspan=2)
totalreserve_ety.place(height=20, width=80, rely=0.59, relx=0.37)





# Buttons
comp_button = tk.Button(comp_output, text="Browse File", activebackground = "pink", activeforeground = "blue", command=lambda: comp_dialog())
comp_button.place(rely=0.91, relx=0.30)

comp_button2 = tk.Button(comp_output, text="Load comp", command=lambda: Load_comp_data())
comp_button2.place(rely=0.91, relx=0)

# The file/file path text
label_comp = ttk.Label(comp_output, text="Empty")
label_comp.place( rely=0.82, relx=0.01)


tv1 = ttk.Treeview(frame1)
tv1.place(relheight=1, relwidth=1) # set the height and width of the widget to 100% of its container (frame1).

treescrolly = tk.Scrollbar(frame1, orient="vertical", command=tv1.yview) # command means update the y axis view of the widget
treescrollx = tk.Scrollbar(frame1, orient="horizontal", command=tv1.xview) # command means update the x axis view of the widget
tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set) # assign the scrollbars to the Treeview Widget
treescrollx.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
treescrolly.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget


def File_dialog():
    """This Function will open the file explorer and assign the chosen file path to label_file"""
    filename = filedialog.askopenfilename(initialdir="/",
                                          title="Select A File",
                                          filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
    label_file["text"] = filename
    return None




def Load_excel_data():
    """If the file selected is valid this will load the file into the Treeview"""
    file_path = label_file["text"]
    try:
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

    except ValueError:
        tk.messagebox.showerror("Information", "The file you have chosen is invalid")
        return None
    except FileNotFoundError:
        tk.messagebox.showerror("Information", f"No such file as {file_path}")
        return None

    clear_data()
    tv1["column"] = list(df.columns)
    tv1["show"] = "headings"
    for column in tv1["columns"]:
        tv1.heading(column, text=column) # let the column heading = column name

    df_rows = df.to_numpy().tolist() # turns the dataframe into a list of lists
    for row in df_rows:
        tv1.insert("", "end", values=row) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
    return None


def clear_data():
    tv1.delete(*tv1.get_children())
    return None



#  volume computation file loading 

vol1 = ttk.Treeview(volumeframe)
vol1.place(relheight=1, relwidth=1) # set the height and width of the widget to 100% of its container (frame1).

treescrollycomp = tk.Scrollbar(volumeframe, orient="vertical", command=vol1.yview) # command means update the y axis view of the widget
treescrollxcomp = tk.Scrollbar(volumeframe, orient="horizontal", command=vol1.xview) # command means update the x axis view of the widget
vol1.configure(xscrollcommand=treescrollxcomp.set, yscrollcommand=treescrollycomp.set) # assign the scrollbars to the Treeview Widget
treescrollxcomp.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
treescrollycomp.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget


def comp_dialog():
    """This Function will open the file explorer and assign the chosen file path to label_file"""
    filename = filedialog.askopenfilename(initialdir="/",
                                          title="Select A File",
                                          filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
    label_comp["text"] = filename
    return None




def Load_comp_data():
    """If the file selected is valid this will load the file into the Treeview"""
    file_path = label_comp["text"]
    try:
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

    except ValueError:
        tk.messagebox.showerror("Information", "The file you have chosen is invalid")
        return None
    except FileNotFoundError:
        tk.messagebox.showerror("Information", f"No such file as {file_path}")
        return None

    clear_data()
    vol1["column"] = list(df.columns)
    vol1["show"] = "headings"
    for column in vol1["columns"]:
        vol1.heading(column, text=column) # let the column heading = column name

    df_rows = df.to_numpy().tolist() # turns the dataframe into a list of lists
    for row in df_rows:
        vol1.insert("", "end", values=row) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
    return None


def clear_data():
    vol1.delete(*vol1.get_children())
    return None


class App(tk.Frame):
    
#    def __init__(self, master):
#         tk.Frame.__init__(self, master, height=42, width=42)
#         self.entry = tk.Entry(self)
#         self.entry.focus()
#         self.entry.pack()
#         self.clear_button = tk.Button(self, text="Clear text", command=self.clear_text)
#         self.clear_button.pack()

#    def clear_text(self):
#        self.entry.delete(0, 'end')

#    def main():
#         root = tk.Tk()
#         App(root).pack(expand=True, fill='both')
#         root.mainloop()

# if __name__ == "__main__":
#     App.main()


    root.mainloop()